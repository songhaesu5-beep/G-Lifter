import random, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(20260808)  # 재현 가능하게 시드 고정

# ---------- 상수 ----------
ZONES = [chr(65+i) for i in range(22)]          # A~V
DEPTHS = list(range(1,6))                        # 1~5
BUNDLE_COUNT = 3                                 # 셀당 bundle 수
ROWS = 5                                         # bundle 내 row 수
COLS = 22                                        # bundle 내 column 수
PER_BUNDLE = ROWS * COLS                         # 110
PER_CELL = BUNDLE_COUNT * PER_BUNDLE             # 330
# 위험률 생성 기준 (행3 셔플링 30%)
#   3번째 행의 모든 차량 중 30%만 무작위로 셔플링 발생
#   → 사람이 출고 불가능하여 G-Lifter 도입이 필요한 차량
SHUFFLE_RATE = 0.30
TARGET_ROW = 3                                   # 셔플링 대상 행 (가운데)
# 위험도 등급은 셔플링 30% 구조의 blocking 분포(11~29) 기반 3분위로 동적 산출

MODELS = ['IONIQ5','EV6','ID.4','Sportage','Tucson','Telluride','Elantra','GV80','G80','Golf','Grandeur']
MFRS = {'IONIQ5':'HYU','EV6':'KIA','ID.4':'VW','Sportage':'KIA','Tucson':'HYU','Telluride':'KIA','Elantra':'HYU','GV80':'GEN','G80':'GEN','Golf':'VW','Grandeur':'HYU'}

# ---------- 번호판 문자 (가~호 35자) ----------
PLATE_CHARS = list('가나다라마거너더러머버서어저고노도로모보소오조구누두루무부수우주하허호')

def gen_plates(count, seed=20260808):
    """XXO XXXX 형식 번호판 count개 중복 없이 생성 (O=한글 35자, X=숫자)"""
    rnd = random.Random(seed)
    plates = set()
    while len(plates) < count:
        plates.add(f'{rnd.randint(0, 99):02d}{rnd.choice(PLATE_CHARS)} {rnd.randint(0, 9999):04d}')
    return list(plates)

# ---------- 실제 GLV2608 명세서 50대 (BL_NO, model, mfr, MTO_MTS, PIO) ----------
REAL = [
 ('HGLV0001','IONIQ5','HYU','MTO','Y'),('HGLV0002','Telluride','KIA','MTS','N'),('HGLV0003','Tucson','HYU','MTS','N'),
 ('VWGL0004','ID.4','VW','MTO','Y'),('HGLV0005','Elantra','HYU','MTS','Y'),
 ('HGLV0006','Tucson','HYU','MTS','N'),('HGLV0007','Tucson','HYU','MTS','N'),('HGLV0008','Tucson','HYU','MTS','N'),('HGLV0009','Tucson','HYU','MTS','N'),('HGLV0010','Tucson','HYU','MTS','N'),
 ('GENG0011','GV80','GEN','MTO','Y'),('GENG0012','G80','GEN','MTS','N'),('HGLV0013','IONIQ5','HYU','MTO','Y'),('KGLV0014','EV6','KIA','MTS','N'),('VWGL0015','Golf','VW','MTO','Y'),
 ('KGLV0016','Sportage','KIA','MTS','N'),('KGLV0017','Sportage','KIA','MTS','N'),('KGLV0018','Sportage','KIA','MTS','N'),('KGLV0019','Sportage','KIA','MTS','N'),('KGLV0020','Sportage','KIA','MTS','N'),
 ('VWGL0021','ID.4','VW','MTS','N'),('VWGL0022','ID.4','VW','MTS','N'),('VWGL0023','ID.4','VW','MTS','N'),('VWGL0024','ID.4','VW','MTS','N'),('VWGL0025','ID.4','VW','MTS','N'),
 ('HGLV0026','Grandeur','HYU','MTO','Y'),('KGLV0027','Telluride','KIA','MTO','Y'),('HGLV0028','Elantra','HYU','MTS','N'),('GENG0029','GV80','GEN','MTO','Y'),('VWGL0030','Golf','VW','MTS','N'),
 ('KGLV0031','EV6','KIA','MTO','Y'),('KGLV0032','EV6','KIA','MTO','Y'),('KGLV0033','EV6','KIA','MTS','N'),('KGLV0034','EV6','KIA','MTS','N'),('KGLV0035','EV6','KIA','MTS','N'),
 ('HGLV0036','IONIQ5','HYU','MTS','N'),('HGLV0037','IONIQ5','HYU','MTS','N'),('HGLV0038','IONIQ5','HYU','MTS','N'),('HGLV0039','IONIQ5','HYU','MTS','N'),('HGLV0040','IONIQ5','HYU','MTS','N'),
 ('GENG0041','G80','GEN','MTO','Y'),('KGLV0042','Telluride','KIA','MTS','N'),('HGLV0043','Tucson','HYU','MTS','N'),('HGLV0044','Tucson','HYU','MTO','Y'),('HGLV0045','Elantra','HYU','MTS','N'),
 ('KGLV0046','Sportage','KIA','MTS','N'),('KGLV0047','Sportage','KIA','MTS','N'),('KGLV0048','Sportage','KIA','MTS','N'),('KGLV0049','Sportage','KIA','MTS','N'),('KGLV0050','Sportage','KIA','MTS','N'),
]

def sim_car(seq):
    model = random.choice(MODELS)
    mto = 'MTO' if random.random() < 0.22 else 'MTS'
    pio = 'Y' if (mto=='MTO' or random.random()<0.08) else 'N'
    return {'bl': f'SIM{seq:05d}', 'model': model, 'mfr': MFRS[model], 'mto': mto, 'pio': pio, 'real': False}

# ---------- 야드 생성: 110셀 × 3 bundle × (5 row × 22 column) = 36,300 슬롯 ----------
cells = []
real_idx = 0
sim_seq = 0
for zone in ZONES:
    for depth in DEPTHS:
        cell = {'id': zone+str(depth), 'zone': zone, 'depth': depth, 'bundles': []}
        for bi in range(BUNDLE_COUNT):
            bundle = []
            for r in range(1, ROWS+1):
                for col in range(1, COLS+1):
                    if real_idx < len(REAL):
                        bl, model, mfr, mto, pio = REAL[real_idx]
                        real_idx += 1
                        car = {'bl': bl, 'model': model, 'mfr': mfr, 'mto': mto, 'pio': pio, 'real': True}
                    else:
                        sim_seq += 1
                        car = sim_car(real_idx + sim_seq)
                    car['row'] = r
                    car['column'] = col
                    car['bundle'] = bi + 1
                    car['pos'] = (r - 1) * COLS + col   # bundle 내 선형 위치 (입구 방향 = 작은 값)
                    bundle.append(car)
            cell['bundles'].append({'cars': bundle, 'blockers': 0, 'shuffle': False})
        cells.append(cell)

# ---------- 공차 처리: 300대 공차 → 야드 수용 가능 차량 36,000대로 고정 ----------
EMPTY_BUDGET = 300
rnd2 = random.Random(20260810)
all_slots = [car for cell in cells for b in cell['bundles'] for car in b['cars']]
rnd2.shuffle(all_slots)
for car in all_slots[:EMPTY_BUDGET]:
    car['empty'] = True
    car['bl'] = car['plate'] = car['model'] = car['mfr'] = ''
    car['mto'] = 'EMPTY'
    car['pio'] = 'N'

# ---------- 위험률 생성 (행3 셔플링 30%) ----------
# 3번째 행(row 3)의 모든 차량(공차 제외) 중 30%만 무작위로 셔플링 발생으로 설정
# → 사람이 출고 불가능하여 G-Lifter 도입이 필요한 차량
# 나머지 행(row 1·2·4·5) 및 행3의 나머지 70%는 인간 출고(Direct)
rnd = random.Random(20260809)
for cell in cells:
    for b in cell['bundles']:
        shuffle_count = 0
        for col in range(1, COLS+1):
            c = b['cars'][(TARGET_ROW-1)*COLS + col - 1]
            if c.get('empty'):
                c['shuffle'] = False
                continue
            if rnd.random() < SHUFFLE_RATE:
                c['shuffle'] = True
                c['mto'] = 'MTO'; c['pio'] = 'Y'; c['target'] = True
                shuffle_count += 1
            else:
                c['shuffle'] = False
        b['blockers'] = shuffle_count

# 셀 통계 (blocking/위험도)
for cell in cells:
    cell['blockers'] = sum(b['blockers'] for b in cell['bundles'])
    cell['delay_min'] = round(cell['blockers'] * 2.3)
    cell['g_after'] = max(2, round(cell['blockers'] * 0.8)) if cell['blockers'] > 0 else 0

# 위험도 등급: 새 분포 기반 3분위 (상위 33% HIGH / 중위 33% MID / 하위 34% OK)
sorted_blockers = sorted(c['blockers'] for c in cells)
n = len(sorted_blockers)
high_thr = sorted_blockers[int(n * 0.67)]   # HIGH 시작 (상위 33%)
mid_thr = sorted_blockers[int(n * 0.34)]    # MID 시작 (중간 33%)
for cell in cells:
    cell['risk'] = 'HIGH' if cell['blockers'] >= high_thr else ('MID' if cell['blockers'] >= mid_thr else 'OK')

# G-Lifter 권고 = blocking 상위 3셀
top3 = sorted([c for c in cells if c['blockers']>0], key=lambda x: -x['blockers'])[:3]
top3_ids = {c['id'] for c in top3}

# ---------- 전체 차량(공차 제외)에 번호판 부여 (중복 없음) ----------
all_cars = [car for cell in cells for b in cell['bundles'] for car in b['cars'] if not car.get('empty')]
plates = gen_plates(len(all_cars))
pi = 0
for car in all_cars:
    car['plate'] = plates[pi]
    pi += 1

# ---------- 엑셀 생성 ----------
wb = Workbook()

# Sheet 1: MASTER (차량 36,000대 + 빈 슬롯 300 = 36,300 그리드)
ws1 = wb.active
ws1.title = 'MASTER_차량'
h1 = ['cell_id','bundle','row','column','plate_no','bl_no','model','mfr','order_type','pio_flag','shuffle_yn']
ws1.append(h1)
header_fill = PatternFill('solid', fgColor='1D4ED8')
for col in range(1, len(h1)+1):
    c = ws1.cell(row=1, column=col)
    c.font = Font(bold=True, color='FFFFFF'); c.fill = header_fill

row_n = 2
for cell in cells:
    for bundle in cell['bundles']:
        for car in bundle['cars']:
            ws1.append([
                cell['id'], car['bundle'], car['row'], car['column'],
                car['plate'], car['bl'], car['model'], car['mfr'], car['mto'], car['pio'],
                'Y' if car.get('shuffle') else 'N'
            ])
            row_n += 1

# Sheet 2: CELL_SUMMARY (110셀 요약)
ws2 = wb.create_sheet('CELL_SUMMARY')
h2 = ['cell_id','zone','depth','bundle_count','bundle_size','capacity','total_vehicles','mto_count','mts_count','target_count','blocking_count','shuffle_count','delay_min','g_lifter_after_min','risk_level','g_lifter_recommend','rank']
ws2.append(h2)
for col in range(1, len(h2)+1):
    c = ws2.cell(row=1, column=col); c.font = Font(bold=True, color='FFFFFF'); c.fill = header_fill

ranked = sorted(cells, key=lambda x: -x['blockers'])
rank_map = {c['id']: i+1 for i, c in enumerate(ranked)}
for cell in cells:
    mto = sum(1 for b in cell['bundles'] for c in b['cars'] if c['mto']=='MTO')
    mts = sum(1 for b in cell['bundles'] for c in b['cars'] if c['mto']=='MTS')
    veh = sum(1 for b in cell['bundles'] for c in b['cars'] if not c.get('empty'))
    ws2.append([
        cell['id'], cell['zone'], cell['depth'], BUNDLE_COUNT, PER_BUNDLE, PER_CELL, veh,
        mto, mts, cell['blockers'], cell['blockers'], cell['blockers'],
        cell['delay_min'], cell['g_after'], cell['risk'],
        'Y' if cell['id'] in top3_ids else 'N', rank_map[cell['id']]
    ])
    # 위험도 셀 색
    risk_color = {'HIGH':'FCA5A5','MID':'FDBA74','OK':'FDE68A'}[cell['risk']]
    ws2.cell(row=ws2.max_row, column=15).fill = PatternFill('solid', fgColor=risk_color)

# Sheet 3: README
ws3 = wb.create_sheet('README')
notes = [
    ['G-Lifter 배치 의사결정 시스템 - 야드 마스터 데이터'],
    [''],
    ['1. 파일 구성'],
    ['  - MASTER_차량 : 실차 36,000대 + 공차 300 (110셀 × 330 슬롯 = 36,300)'],
    ['  - CELL_SUMMARY : 셀별 blocking/위험도/우선순위 요약 (110행)'],
    [''],
    ['1-1. MASTER_차량 컬럼'],
    ['  - cell_id / bundle / row / column / plate_no / bl_no / model / mfr / order_type / pio_flag / shuffle_yn'],
    ['  - bundle 내 선형 위치 pos = (row-1)*22 + column'],
    ['  - 공차: order_type=EMPTY, plate_no/bl_no/model/mfr 공란'],
    ['  - shuffle_yn: 행3 셔플링 발생 차량(사람 출고 불가, G-Lifter 필요) = Y'],
    [''],
    ['2. 야드 구조'],
    ['  - 22 Zone(A~V) × 5 Depth = 110셀'],
    ['  - 셀 1개 = 3 bundle × (5 row × 22 column) = 330 슬롯'],
    ['  - 수용 규모: 36,300 슬롯 중 실차 36,000대 + 공차 300 (고정)'],
    [''],
    ['3. 셔플링 발생 기준 (위험률 생성)'],
    ['  - 3번째 행(row 3)의 차량(공차 제외) 중 30%만 무작위로 셔플링 발생'],
    ['  - 셔플링 차량 = 사람이 출고 불가능 → G-Lifter 도입 필요 (order_type=MTO, 대상 지정)'],
    ['  - 나머지 행(1·2·4·5) 및 행3의 나머지 70%는 인간 출고(Direct)'],
    ['  - bundle blocking = 셔플링 발생 차량 수 / 셀 blocking = 3개 bundle 합계'],
    [''],
    ['4. 위험도 기준'],
    ['  - 셔플링 30% 구조에서 셀별 blocking은 11~29대 분포'],
    ['  - 새 분포 기반 3분위: 상위 33% HIGH / 중위 33% MID / 하위 34% OK (동적 임계값)'],
    ['  - G-Lifter 권고: blocking 상위 3셀'],
    [''],
    ['5. 시드 고정'],
    ['  - 시드 20260808(야드) / 20260809(셔플링) / 20260810(공차) → 동일 파일 재현 가능'],
    ['  - GLV2608 실데이터 50대가 셀 A1부터 배치됨'],
]
for r in notes:
    ws3.append(r)
ws3.column_dimensions['A'].width = 90

# 컬럼 너비
for ws, widths in [(ws1, [9,7,5,7,12,12,10,6,10,9,10]), (ws2, [8,5,6,12,11,9,13,9,9,12,13,13,9,16,10,15,5])]:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

out = 'G-Lifter_raw_data.xlsx'
wb.save(out)

print('생성 완료:', out)
all_cars_flat = [car for cell in cells for bundle in cell['bundles'] for car in bundle['cars']]
all_plates = [c['plate'] for c in all_cars_flat if not c.get('empty')]
empty_count = sum(1 for c in all_cars_flat if c.get('empty'))
shuffle_cars = sum(1 for c in all_cars_flat if c.get('shuffle'))
row3_vehicles = sum(1 for c in all_cars_flat if c['row'] == TARGET_ROW and not c.get('empty'))
print('실차 수:', len(all_plates), '| 공차:', empty_count)
print('셀 수:', len(cells), '| bundle 수:', sum(len(c['bundles']) for c in cells))
print('셔플링 차량(행3 30%):', shuffle_cars, '(', round(shuffle_cars/row3_vehicles*100), '% of 행3 실차 )')
print('번호판 중복 확인:', 'OK' if len(set(all_plates)) == len(all_plates) else 'FAIL')
print('위험도 분포(3분위): HIGH', sum(1 for c in cells if c['risk']=='HIGH'),
      '/ MID', sum(1 for c in cells if c['risk']=='MID'),
      '/ OK', sum(1 for c in cells if c['risk']=='OK'), '| 임계값:', mid_thr, '~', high_thr)
print('G-Lifter 권고 셀:', ', '.join(sorted(top3_ids)))
print('실데이터 배치:', real_idx, '대')
