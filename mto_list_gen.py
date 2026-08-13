import csv, random
from openpyxl import load_workbook

random.seed(20260808)  # 재현 가능하게 시드 고정

SRC = 'G-Lifter_raw_data2.xlsx'
OUT = 'mto_list.csv'
OUT2 = 'list2.csv'
OUT3 = 'list3.csv'
SAMPLE_N = 10
SAMPLE2_N = 5000
SAMPLE3_N = 5000
SEED3 = 20260811

wb = load_workbook(SRC, read_only=True)
ws = wb['MASTER_차량']
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[4]]  # 빈 슬롯(plate_no 공란) 제외
wb.close()

selected = random.sample(rows, SAMPLE_N)
selected2 = random.sample(rows, SAMPLE2_N)
rnd3 = random.Random(SEED3)   # 별도 시드로 기존 출력 영향 없음
selected3 = rnd3.sample(rows, SAMPLE3_N)

with open(OUT, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerow(['plate_no'])
    writer.writerows([[r[4]] for r in selected])

with open(OUT2, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerow(['plate_no'])
    writer.writerows([[r[4]] for r in selected2])

with open(OUT3, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerow(['plate_no'])
    writer.writerows([[r[4]] for r in selected3])

print('생성 완료:', OUT, '|', OUT2, '|', OUT3)
print('전체 차량:', len(rows), '| 추출:', len(selected), '/', len(selected2), '/', len(selected3), '(시드', SEED3, ')')
print(OUT3, '첫 10개:')
for r in selected3[:10]:
    print(' ', r[4])
